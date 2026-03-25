"""
Automação - Check-in AP: Aba "AP By GL"
========================================
Este script consolida os 6 relatórios brutos do RLM (AP Purchase Journal By GL),
aplica todas as transformações e lookups necessários, e gera o arquivo tratado
com as 46 colunas da aba "AP By GL".

USO:
    python automacao_checkin_ap.py

ESTRUTURA DE PASTAS ESPERADA:
    ./relatorios_brutos/     → Colocar os 6 arquivos .xlsx do RLM aqui
    ./kyriba_ca.xlsx         → Arquivo exportado do Kyriba (C.A.)
    ./output/                → Arquivo de saída gerado aqui

CONFIGURAÇÃO:
    Ajuste os caminhos nas variáveis abaixo conforme necessário.
"""

import pandas as pd
import numpy as np
import os
import glob
import warnings
from datetime import datetime, timedelta

warnings.filterwarnings('ignore')

# ============================================================
# CONFIGURAÇÃO DE CAMINHOS
# ============================================================
PASTA_RELATORIOS = "./relatorios_brutos"       # Pasta com os 6 relatórios brutos
ARQUIVO_KYRIBA   = "./kyriba_ca.xlsx"          # Arquivo Kyriba (C.A.)
PASTA_SAIDA      = "./output"                  # Pasta de saída
NOME_SAIDA       = "Check_in_AP_tratado.xlsx"  # Nome do arquivo de saída

# ============================================================
# TABELAS DE REFERÊNCIA INTERNAS (FIXAS)
# ============================================================

# --- Currency (de/para) ---
CURRENCY_MAP = {
    'E': 'EUR',
    '$': 'USD',
    'G': 'GBP',
    'S': 'CHF',
    'B': 'BRL',
    'T': 'TRY',
}

# --- Transfer Type ---
# Cada linha: (Company, Currency, BranchCountry, Código, Código2, TransferType, OtherTransferType, KyribaDebitAccount)
TRANSFER_TYPE_DATA = [
    (1, 'USD', 'US', '1USDUS', '1USD', 'FEDW', 'ITRF', '2CITPLAUSD'),
    (2, 'USD', 'NL', '2USDNL', '2USD', 'DTRF', 'ITRF', '2CITSBICHUSD'),
    (2, 'EUR', 'NL', '2EURNL', '2EUR', 'SEPD', 'ITRF', '2CITSBIEUR'),
    (2, 'GBP', 'GB', '2GBPGB', '2GBP', 'FAST', 'ITRF', '2CITSBIGBP'),
    (2, 'CHF', 'CH', '2CHFCH', '2CHF', 'DTRF', 'ITRF', '2CITSBICHCHF'),
    (3, 'GBP', 'GB', '3GBPGB', '3GBP', 'FAST', 'ITRF', '2CITSBUKGBP'),
    (4, 'EUR', 'FR', '4EURFR', '4EUR', 'SEPD', 'ITRF', '2CITSBFEUR'),
    (6, 'EUR', 'PT', '6EURPT', '6EUR', 'SEPD', 'ITRF', '2CITSBPTEUR'),
]

# Dicionários para lookups rápidos
# Código (ex: "1USDUS") → Transfer Type (ex: "FEDW")
TT_BY_CODIGO = {row[3]: row[5] for row in TRANSFER_TYPE_DATA}
# Código2 (ex: "1USD") → Kyriba Debit Account (ex: "2CITPLAUSD")
TT_DEBIT_BY_CODIGO2 = {row[4]: row[7] for row in TRANSFER_TYPE_DATA}
# Company → Kyriba Debit Account (fallback)
TT_DEBIT_BY_COMPANY = {row[0]: row[7] for row in TRANSFER_TYPE_DATA}

# --- Validação ---
# Tipo 1: lookup por Vendor # (coluna E)
# Tipo 2: lookup por SUFFIX1 (coluna A)
VALIDACAO_BY_VENDOR = {
    '001353': 'NOT PAY - CCC',
    '000044': 'VALIDATE - Lease',
    '000290': 'VALIDATE - Lease',
    '001181': 'VALIDATE - Lease',
    '000848': 'VALIDATE - Lease',
    '001476': 'VALIDATE - Lease',
    '001998': 'VALIDATE - Lease',
    '001880': 'VALIDATE - Lease',
    '001593': 'VALIDATE - Lease',
    '001528': 'VALIDATE - Lease',
    '001779': 'VALIDATE - Lease',
    '001581': 'VALIDATE - Lease',
    '001941': 'VALIDATE - Lease',
    '002089': 'VALIDATE - Lease',
    '003227': 'VALIDATE - Lease',
    '003288': 'VALIDATE - Lease',
    '003830': 'VALIDATE - Lease',
    '001735': 'NOT PAY - PR',
    '001734': 'NOT PAY - PR',
    '001733': 'NOT PAY - PR',
    '001478': 'NOT PAY - PR',
    '001817': 'NOT PAY - PR',
    '002069': 'NOT PAY - PR',
    '003090': 'NOT PAY - PR',
    '002017': 'VALIDATE - WH AGENT ',
    '001708': 'VALIDATE - WH AGENT ',
    '002072': 'VALIDATE - WH AGENT ',
    '003672': 'VALIDATE - WH AGENT ',
    '003699': 'VALIDATE - WH AGENT ',
    '003107': 'VALIDATE - WH AGENT ',
    '003172': 'VALIDATE - WH AGENT ',
    '001871': 'VALIDATE - PAYROLL',
    '003166': 'VALIDATE - PAYROLL',
    '003006': 'VALIDATE - PAYROLL',
    '003007': 'VALIDATE - PAYROLL',
    '003008': 'VALIDATE - PAYROLL',
    '003009': 'VALIDATE - PAYROLL',
    '003010': 'VALIDATE - PAYROLL',
    '003201': 'VALIDATE - PAYROLL',
    '003220': 'VALIDATE - PAYROLL',
    '001352': 'VALIDATE - OZAN/RAFA',
    '001377': 'VALIDATE - OZAN/RAFA',
    '001351': 'VALIDATE - OZAN/RAFA',
    '003338': 'HOLD',
    '000855': 'NOT PAY - Estornar',
    '003265': 'NOT PAY - Estornar',
    '001897': 'NOT PAY - Estornar',
    '003274': 'HOLD',
    '000442': 'NOT PAY - Shopify (CCC)',
}

VALIDACAO_BY_SUFFIX = {
    115003: 'NOT PAY - PP',
    114003: 'NOT PAY - WH',
    114004: 'NOT PAY - WH',
    521997: 'VALIDATE - DUTIES',
    114006: 'VALIDATE - DUTIES',
    114010: 'VALIDATE - Consumível',
}


# ============================================================
# FUNÇÕES
# ============================================================

def carregar_kyriba(caminho):
    """Carrega o arquivo Kyriba (C.A.) e retorna dicionários de lookup."""
    print(f"  Carregando Kyriba (C.A.): {caminho}")
    df = pd.read_excel(caminho)

    # Identifica colunas - baseado na estrutura conhecida:
    # Col B (Name 2) = Vendor # | Col C (Code) = Credit Account | Col K (Country) = Vendor Bank Country
    cols = df.columns.tolist()

    # Procura as colunas por nome ou posição
    col_vendor = cols[1]   # Name 2 (Vendor #)
    col_code   = cols[2]   # Code (Credit Account)
    col_country = cols[10] # Country (Vendor Bank Country)

    df[col_vendor] = df[col_vendor].astype(str).str.strip().str.zfill(6)

    # Vendor # → Country
    kyriba_country = df.set_index(col_vendor)[col_country].to_dict()
    # Vendor # → Code (Credit Account)
    kyriba_credit = df.set_index(col_vendor)[col_code].to_dict()

    print(f"    → {len(kyriba_country)} vendors carregados")
    return kyriba_country, kyriba_credit


def carregar_relatorios(pasta):
    """Carrega todos os .xlsx da pasta, consolida e limpa subtotais."""
    arquivos = glob.glob(os.path.join(pasta, "*.xlsx"))
    if not arquivos:
        raise FileNotFoundError(f"Nenhum arquivo .xlsx encontrado em: {pasta}")

    print(f"  Encontrados {len(arquivos)} relatórios:")
    dfs = []
    for arq in sorted(arquivos):
        print(f"    → {os.path.basename(arq)}")
        df = pd.read_excel(arq)
        dfs.append(df)

    df_all = pd.concat(dfs, ignore_index=True)
    linhas_antes = len(df_all)

    # Remove linhas de subtotal/total: onde Company # (coluna A) está vazio/NaN
    col_company = df_all.columns[0]  # "Company #"
    df_all = df_all[df_all[col_company].notna() & (df_all[col_company] != '')]

    # Garante que Company # é numérico (remove textos residuais)
    df_all = df_all[pd.to_numeric(df_all[col_company], errors='coerce').notna()]

    linhas_depois = len(df_all)
    print(f"  Consolidado: {linhas_antes} → {linhas_depois} linhas ({linhas_antes - linhas_depois} subtotais removidos)")
    return df_all


def calcular_colunas(df, kyriba_country, kyriba_credit):
    """Aplica todas as 17 colunas calculadas."""
    print("  Aplicando colunas calculadas...")

    # Garante tipos corretos
    df['Company #'] = df['Company #'].astype(int)
    # Vendor # deve ser string com 6 dígitos e zeros à esquerda (ex: "000351")
    df['Vendor #'] = pd.to_numeric(df['Vendor #'], errors='coerce').fillna(0).astype(int).astype(str).str.zfill(6)
    df['Vendor Curr'] = df['Vendor Curr'].astype(str).str.strip()
    df['Exchange Rate'] = pd.to_numeric(df['Exchange Rate'], errors='coerce').fillna(1)
    df['Amount($)'] = pd.to_numeric(df['Amount($)'], errors='coerce').fillna(0)

    # Converte SUFFIX1 para numérico onde possível
    df['SUFFIX1 '] = pd.to_numeric(df['SUFFIX1 '], errors='coerce')

    # --- Original Amount ---
    df['Original Amount'] = df['Amount($)'] * df['Exchange Rate']

    # --- Currency ---
    df['Currency'] = df['Vendor Curr'].map(CURRENCY_MAP).fillna(df['Vendor Curr'])

    # --- Vendor Bank Country ---
    df['Vendor Bank Country'] = df['Vendor #'].map(kyriba_country).fillna('Não Cadastrado')

    # --- Credit Account (Kyriba) ---
    df['Credit Account (Kyriba)'] = df['Vendor #'].map(kyriba_credit).fillna('Não Cadastrado')

    # --- Código ---
    df['Código'] = (
        df['Company #'].astype(str)
        + df['Currency']
        + df['Vendor Bank Country']
    )

    # --- Debit Account (Kyriba) ---
    # VLOOKUP(LEFT(Código,4), TransferType!E:H, 4) fallback VLOOKUP(Company#, TransferType!A:H, 8)
    def get_debit_account(row):
        codigo_left4 = str(row['Código'])[:4]
        result = TT_DEBIT_BY_CODIGO2.get(codigo_left4)
        if result:
            return result
        return TT_DEBIT_BY_COMPANY.get(row['Company #'], '')
    df['Debit Account (Kyriba)'] = df.apply(get_debit_account, axis=1)

    # --- Transfer Type ---
    # IF(CreditAccount="Não Cadastrado","Não Cadastrado", IFERROR(VLOOKUP(Código,TT!D:F,3),"ITRF"))
    def get_transfer_type(row):
        if row['Credit Account (Kyriba)'] == 'Não Cadastrado':
            return 'Não Cadastrado'
        result = TT_BY_CODIGO.get(row['Código'])
        return result if result else 'ITRF'
    df['Transfer Type'] = df.apply(get_transfer_type, axis=1)

    # --- Reason ---
    df['Reason'] = 'FARM PYM >' + df['Vendor Name'].astype(str)

    # --- AP Type ---
    df['AP Type'] = df['General Ledger Description'].astype(str).str.lower().apply(
        lambda x: 'Produtivo' if 'inventory' in x else 'Consumível'
    )

    # --- GL ---
    df['GL'] = df.apply(
        lambda row: row['General Ledger Description'] if row['AP Type'] == 'Produtivo' else 'Expense',
        axis=1
    )

    # --- Status ---
    # Paid Date > 0 → PAID, senão UNPAID
    def get_status(paid_date):
        if pd.isna(paid_date):
            return 'UNPAID'
        if isinstance(paid_date, datetime):
            return 'PAID'
        try:
            val = float(paid_date)
            return 'PAID' if val > 0 else 'UNPAID'
        except (ValueError, TypeError):
            return 'PAID' if paid_date else 'UNPAID'
    df['Status'] = df['Paid Date'].apply(get_status)

    # --- Year / Month / Week ---
    def parse_date_safe(val):
        if pd.isna(val):
            return pd.NaT
        if isinstance(val, datetime):
            return val
        try:
            return pd.to_datetime(val)
        except:
            return pd.NaT

    paid_dates_parsed = df['Paid Date'].apply(parse_date_safe)

    df['Year'] = df.apply(
        lambda row: paid_dates_parsed[row.name].year if row['Status'] == 'PAID' and pd.notna(paid_dates_parsed[row.name]) else 'UNPAID',
        axis=1
    )
    df['Month'] = df.apply(
        lambda row: paid_dates_parsed[row.name].month if row['Status'] == 'PAID' and pd.notna(paid_dates_parsed[row.name]) else 'UNPAID',
        axis=1
    )
    df['Week'] = df.apply(
        lambda row: paid_dates_parsed[row.name].isocalendar()[1] if row['Status'] == 'PAID' and pd.notna(paid_dates_parsed[row.name]) else 'UNPAID',
        axis=1
    )

    # --- Validação ---
    def get_validacao(row):
        if row['Status'] == 'PAID':
            return 'PAID'
        if row['Status'] == 'UNPAID':
            vendor = str(row['Vendor #']).strip()
            result = VALIDACAO_BY_VENDOR.get(vendor)
            if result:
                return result
            suffix = row['SUFFIX1 ']
            if pd.notna(suffix):
                try:
                    suffix_int = int(suffix)
                    result = VALIDACAO_BY_SUFFIX.get(suffix_int)
                    if result:
                        return result
                except (ValueError, TypeError):
                    pass
            return '-'
        return '-'
    df['Validação'] = df.apply(get_validacao, axis=1)

    # --- Expt Pymt ---
    # Lógica: próxima sexta-feira para itens UNPAID (pagamento às sextas de tudo aprovado até a sexta anterior)
    def get_expt_pymt(row):
        if row['Status'] == 'PAID':
            return 'PAID'
        today = datetime.today()
        days_until_friday = (4 - today.weekday()) % 7
        if days_until_friday == 0:
            days_until_friday = 7
        next_friday = today + timedelta(days=days_until_friday)
        return next_friday
    df['Expt Pymt'] = df.apply(get_expt_pymt, axis=1)

    # --- OBS ---
    df['OBS'] = None

    print("    → 17 colunas calculadas aplicadas com sucesso")
    return df


def ordenar_colunas(df):
    """Garante a ordem final das 46 colunas."""
    colunas_originais = [
        'Company #', 'Voucher #', 'SUFFIX1 ', 'MAIN ', 'DEPT ',
        'General Ledger Description', 'Season', 'PO #', 'Shipment #',
        'House/Airway Bill #', 'Vendor #', 'Vendor Name', 'Business Type',
        'Entry Date', 'Batch Date', 'Purchase Period', 'Invoice Date',
        'Due Date', 'Paid Date', 'DSO', 'Days Late', 'Check #',
        'Invoice #', 'Description', 'Summ Cost', 'Vendor Curr',
        'Exchange Rate', 'Amount($)', 'Voucher Notes',
    ]
    colunas_calculadas = [
        'Original Amount', 'Currency', 'Vendor Bank Country',
        'Credit Account (Kyriba)', 'Código', 'Debit Account (Kyriba)',
        'Transfer Type', 'Reason', 'GL', 'AP Type', 'Status',
        'Year', 'Month', 'Week', 'Validação', 'Expt Pymt', 'OBS',
    ]
    ordem_final = colunas_originais + colunas_calculadas

    # Mantém apenas colunas que existem
    colunas_presentes = [c for c in ordem_final if c in df.columns]
    return df[colunas_presentes]


def formatar_excel(caminho_saida, df):
    """Salva o DataFrame como .xlsx com formatação profissional."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "AP By GL"

    # Header
    header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    # Escrita dos headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Escrita dos dados
    data_font = Font(name='Arial', size=9)
    for row_idx, row_data in enumerate(df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(value, (np.integer, np.int64)):
                cell.value = int(value)
            elif isinstance(value, (np.floating, np.float64)):
                cell.value = float(value)
            elif isinstance(value, np.bool_):
                cell.value = bool(value)
            elif pd.isna(value) if not isinstance(value, str) else False:
                cell.value = None
            else:
                cell.value = value
            cell.font = data_font
            cell.border = thin_border

    # Formatação de colunas monetárias
    money_cols = ['Amount($)', 'Original Amount']
    for col_name in money_cols:
        if col_name in df.columns:
            col_idx = list(df.columns).index(col_name) + 1
            for row in range(2, len(df) + 2):
                ws.cell(row=row, column=col_idx).number_format = '#,##0.00'

    # Auto-ajuste de largura (com limite)
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(df.columns[col_idx - 1]))
        for row in range(2, min(len(df) + 2, 102)):  # Amostra de 100 linhas
            val = ws.cell(row=row, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

    # Filtro automático
    ws.auto_filter.ref = ws.dimensions

    # Congelar painel (header fixo)
    ws.freeze_panes = 'A2'

    wb.save(caminho_saida)
    print(f"    → Arquivo salvo: {caminho_saida}")


# ============================================================
# EXECUÇÃO PRINCIPAL
# ============================================================
def main():
    print("=" * 60)
    print("  AUTOMAÇÃO CHECK-IN AP - Aba AP By GL")
    print("=" * 60)

    # 1. Carrega Kyriba
    print("\n[1/4] Carregando Kyriba (C.A.)...")
    kyriba_country, kyriba_credit = carregar_kyriba(ARQUIVO_KYRIBA)

    # 2. Carrega e consolida relatórios
    print("\n[2/4] Carregando e consolidando relatórios brutos...")
    df = carregar_relatorios(PASTA_RELATORIOS)

    # 3. Aplica colunas calculadas
    print("\n[3/4] Calculando colunas...")
    df = calcular_colunas(df, kyriba_country, kyriba_credit)
    df = ordenar_colunas(df)

    # 4. Salva resultado
    print("\n[4/4] Salvando arquivo de saída...")
    os.makedirs(PASTA_SAIDA, exist_ok=True)
    caminho_saida = os.path.join(PASTA_SAIDA, NOME_SAIDA)
    formatar_excel(caminho_saida, df)

    print(f"\n{'=' * 60}")
    print(f"  CONCLUÍDO! {len(df)} linhas processadas")
    print(f"  Saída: {caminho_saida}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
