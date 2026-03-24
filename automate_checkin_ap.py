"""
=============================================================================
AUTOMAÇÃO CHECK-IN AP — FARM Global Finance
=============================================================================
Script para automatizar o tratamento da planilha CheckinAP.
Lê dados brutos do ERP (RLM) e aplica todas as transformações da aba "AP by GL".

Uso:
    python automate_checkin_ap.py --bases ./Bases --ref ./CheckinAP_Ref.xlsx --output ./CheckinAP_Automatizado.xlsx

Dependências:
    pip install pandas openpyxl numpy
=============================================================================
"""

import argparse
import glob
import os
import sys
from datetime import datetime, timedelta

import numpy as np
import pandas as pd


# ═══════════════════════════════════════════════════════════════════════════════
# 1. CONFIGURAÇÃO E CONSTANTES
# ═══════════════════════════════════════════════════════════════════════════════

RAW_COLUMNS = [
    'Company #', 'Voucher #', 'SUFFIX1 ', 'MAIN ', 'DEPT ',
    'General Ledger Description', 'Season', 'PO #', 'Shipment #',
    'House/Airway Bill #', 'Vendor #', 'Vendor Name', 'Business Type',
    'Entry Date', 'Batch Date', 'Purchase Period', 'Invoice Date',
    'Due Date', 'Paid Date', 'DSO', 'Days Late', 'Check #',
    'Invoice #', 'Description', 'Summ Cost', 'Vendor Curr',
    'Exchange Rate', 'Amount($)', 'Voucher Notes'
]

CALC_COLUMNS = [
    'Original Amount', 'Currency', 'Vendor Bank Country',
    'Credit Account (Kyriba)', 'Código', 'Debit Account (Kyriba)',
    'Transfer Type', 'Reason', 'GL', 'AP Type', 'Status',
    'Year', 'Month', 'Week', 'Validação', 'Expt Pymt', 'OBS'
]


# ═══════════════════════════════════════════════════════════════════════════════
# 2. LEITURA E CONSOLIDAÇÃO DAS BASES
# ═══════════════════════════════════════════════════════════════════════════════

def load_and_consolidate_bases(bases_path):
    """
    Lê todos os arquivos Excel de uma pasta (exportações do ERP por company),
    consolida em um único DataFrame e remove linhas de subtotal/total.
    """
    files = glob.glob(os.path.join(bases_path, "*.xlsx")) + \
            glob.glob(os.path.join(bases_path, "*.xls"))

    if not files:
        print(f"[ERRO] Nenhum arquivo encontrado em: {bases_path}")
        sys.exit(1)

    frames = []
    for f in sorted(files):
        print(f"  Lendo: {os.path.basename(f)}")
        try:
            df = pd.read_excel(f, sheet_name=0)
            frames.append(df)
        except Exception as e:
            print(f"  [AVISO] Erro ao ler {os.path.basename(f)}: {e}")

    if not frames:
        print("[ERRO] Nenhum arquivo pôde ser lido.")
        sys.exit(1)

    consolidated = pd.concat(frames, ignore_index=True)
    initial_rows = len(consolidated)

    # Filtrar linhas em branco na coluna A (Company #) — subtotais e totais
    first_col = consolidated.columns[0]
    consolidated = consolidated[consolidated[first_col].notna()].copy()
    consolidated = consolidated[
        pd.to_numeric(consolidated[first_col], errors='coerce').notna()
    ].copy()

    removed = initial_rows - len(consolidated)
    print(f"  Consolidado: {len(consolidated)} linhas ({removed} subtotais removidos)")
    return consolidated


def load_single_file(filepath):
    """
    Lê um único arquivo CheckinAP e retorna apenas colunas brutas.
    """
    print(f"  Lendo: {os.path.basename(filepath)}")
    df = pd.read_excel(filepath, sheet_name="AP By GL")

    cols_to_drop = [c for c in df.columns if c in CALC_COLUMNS]
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)
        print(f"  Removidas {len(cols_to_drop)} colunas calculadas existentes")

    return df


# ═══════════════════════════════════════════════════════════════════════════════
# 3. CARREGAMENTO DAS TABELAS DE REFERÊNCIA
# ═══════════════════════════════════════════════════════════════════════════════

def load_reference_tables(ref_path):
    """
    Carrega todas as abas de referência do arquivo de referência.
    """
    refs = {}
    print("\n[2] Carregando tabelas de referência...")

    # ── Currency ──
    refs['currency_df'] = pd.read_excel(ref_path, sheet_name="Currency")
    refs['currency_map'] = dict(zip(
        refs['currency_df']['Vendor Curr'].astype(str).str.strip(),
        refs['currency_df']['Currency'].astype(str).str.strip()
    ))
    print(f"  Currency: {len(refs['currency_map'])} mapeamentos")

    # ── Kyriba (C.A.) ──
    refs['kyriba_ca_df'] = pd.read_excel(ref_path, sheet_name="Kyriba (C.A.)")
    kyriba = refs['kyriba_ca_df'].copy()
    kyriba['vendor_key'] = kyriba['Name 2'].astype(str).str.strip().str.lstrip('0')
    refs['kyriba_credit_map'] = dict(zip(kyriba['vendor_key'], kyriba['Code']))
    refs['kyriba_country_map'] = dict(zip(kyriba['vendor_key'], kyriba['Country']))
    print(f"  Kyriba (C.A.): {len(kyriba)} registros")

    # ── Transfer Type ──
    refs['transfer_type_df'] = pd.read_excel(ref_path, sheet_name="Transfer Type")
    tt = refs['transfer_type_df']
    refs['tt_transfer_map'] = dict(zip(tt['Código'].astype(str).str.strip(), tt['Transfer Type'].astype(str).str.strip()))
    refs['tt_debit_map'] = dict(zip(tt['Código'].astype(str).str.strip(), tt['Kyriba Debit Account'].astype(str).str.strip()))
    refs['tt_other_map'] = dict(zip(tt['Código'].astype(str).str.strip(), tt['Other Transfer Type'].astype(str).str.strip()))
    refs['tt_transfer_map2'] = dict(zip(tt['Código 2'].astype(str).str.strip(), tt['Transfer Type'].astype(str).str.strip()))
    refs['tt_debit_map2'] = dict(zip(tt['Código 2'].astype(str).str.strip(), tt['Kyriba Debit Account'].astype(str).str.strip()))
    print(f"  Transfer Type: {len(tt)} registros")

    # ── Validação ──
    refs['validacao_df'] = pd.read_excel(ref_path, sheet_name="Validação")
    val = refs['validacao_df']

    # Mapa por Vendor # (coluna index 4 = "General Ledger Description.1")
    vendor_val = val[val.iloc[:, 4].notna()].copy()
    refs['val_vendor_map'] = dict(zip(
        vendor_val.iloc[:, 4].astype(float).astype(int).astype(str),
        vendor_val['Validação'].astype(str)
    ))

    # Mapa por SUFFIX1 (para linhas de GL específico)
    suffix_val = val[(val.iloc[:, 0] != '-') & (val.iloc[:, 0].notna())].copy()
    suffix_keys = {}
    for _, row in suffix_val.iterrows():
        suffix = str(row.iloc[0]).strip()
        main_val = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
        dept_val = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
        gl_val = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
        # Chave composta: SUFFIX|MAIN|DEPT|GL
        key = f"{suffix}|{main_val}|{dept_val}|{gl_val}"
        suffix_keys[key] = str(row['Validação'])
        # Chave simplificada: apenas SUFFIX
        suffix_keys[suffix] = str(row['Validação'])
    refs['val_suffix_map'] = suffix_keys
    print(f"  Validação: {len(refs['val_vendor_map'])} por Vendor, {len(suffix_keys)} por SUFFIX")

    return refs


# ═══════════════════════════════════════════════════════════════════════════════
# 4. TRANSFORMAÇÕES
# ═══════════════════════════════════════════════════════════════════════════════

def apply_transformations(df, refs):
    """
    Aplica todas as transformações das colunas AD em diante.
    """
    print("\n[3] Aplicando transformações...")

    # Preparação de tipos
    df['Amount($)'] = pd.to_numeric(df['Amount($)'], errors='coerce').fillna(0)
    df['Exchange Rate'] = pd.to_numeric(df['Exchange Rate'], errors='coerce').fillna(1)
    df['Company #'] = pd.to_numeric(df['Company #'], errors='coerce').astype('Int64')
    df['Vendor #'] = df['Vendor #'].astype(str).str.strip().str.lstrip('0')

    df['Paid Date'] = pd.to_datetime(df['Paid Date'], errors='coerce')
    df['Due Date'] = pd.to_datetime(df['Due Date'], errors='coerce')

    # ── Original Amount ──
    df['Original Amount'] = df['Amount($)'] * df['Exchange Rate']
    print("  ✓ Original Amount = Amount($) × Exchange Rate")

    # ── Currency ──
    df['Currency'] = df['Vendor Curr'].astype(str).str.strip().map(
        refs['currency_map']
    ).fillna(df['Vendor Curr'])
    print("  ✓ Currency (PROCV → aba Currency)")

    # ── Vendor Bank Country ──
    df['Vendor Bank Country'] = df['Vendor #'].map(
        refs['kyriba_country_map']
    ).fillna('Não Cadastrado')
    print("  ✓ Vendor Bank Country (PROCV → Kyriba C.A.)")

    # ── Credit Account (Kyriba) ──
    df['Credit Account (Kyriba)'] = df['Vendor #'].map(
        refs['kyriba_credit_map']
    ).fillna('Não Cadastrado')
    print("  ✓ Credit Account (PROCV → Kyriba C.A.)")

    # ── Código ──
    df['Código'] = (
        df['Company #'].astype(str) +
        df['Currency'].astype(str) +
        df['Vendor Bank Country'].astype(str)
    )
    print("  ✓ Código = Company# + Currency + Country")

    # ── Debit Account (Kyriba) ──
    df['Debit Account (Kyriba)'] = df['Código'].map(refs['tt_debit_map'])
    codigo2 = df['Company #'].astype(str) + df['Currency'].astype(str)
    mask = df['Debit Account (Kyriba)'].isna()
    df.loc[mask, 'Debit Account (Kyriba)'] = codigo2[mask].map(refs['tt_debit_map2'])
    df['Debit Account (Kyriba)'] = df['Debit Account (Kyriba)'].fillna('Não Cadastrado')
    print("  ✓ Debit Account (PROCV → Transfer Type)")

    # ── Transfer Type ──
    df['Transfer Type'] = df['Código'].map(refs['tt_transfer_map'])
    mask_no_credit = df['Credit Account (Kyriba)'] == 'Não Cadastrado'
    df.loc[mask_no_credit, 'Transfer Type'] = 'Não Cadastrado'
    mask_na = df['Transfer Type'].isna()
    df.loc[mask_na, 'Transfer Type'] = codigo2[mask_na].map(refs['tt_transfer_map2'])
    df['Transfer Type'] = df['Transfer Type'].fillna('Não Cadastrado')
    print("  ✓ Transfer Type")

    # ── Reason ──
    df['Reason'] = 'FARM PYM >' + df['Vendor Name'].astype(str)
    print("  ✓ Reason = 'FARM PYM >' + Vendor Name")

    # ── AP Type ──
    gl_lower = df['General Ledger Description'].astype(str).str.lower()
    df['AP Type'] = np.where(
        gl_lower.str.contains('inventory', na=False),
        'Produtivo', 'Consumível'
    )
    print("  ✓ AP Type (inventory → Produtivo, else → Consumível)")

    # ── GL ──
    df['GL'] = np.where(
        df['AP Type'] == 'Produtivo',
        df['General Ledger Description'], 'Expense'
    )
    print("  ✓ GL (Produtivo → GL Description, Consumível → Expense)")

    # ── Status ──
    df['Status'] = np.where(df['Paid Date'].notna(), 'PAID', 'UNPAID')
    print("  ✓ Status (Paid Date preenchida → PAID)")

    # ── Year / Month / Week ──
    df['Year'] = np.where(df['Status'] == 'PAID', df['Paid Date'].dt.year, 'UNPAID')
    df['Month'] = np.where(df['Status'] == 'PAID', df['Paid Date'].dt.month, 'UNPAID')
    week_vals = df['Paid Date'].dt.isocalendar().week.values
    df['Week'] = np.where(df['Status'] == 'PAID', week_vals, 'UNPAID')
    print("  ✓ Year / Month / Week")

    # ── Validação ──
    def calc_validacao(row):
        if row['Status'] == 'PAID':
            return 'PAID'
        vendor = str(row['Vendor #']).strip()
        if vendor in refs['val_vendor_map']:
            return refs['val_vendor_map'][vendor]
        suffix_col = 'SUFFIX1 ' if 'SUFFIX1 ' in row.index else 'SUFFIX1'
        suffix = str(row.get(suffix_col, '')).strip()
        main_v = str(row.get('MAIN ', row.get('MAIN', ''))).strip()
        dept_v = str(row.get('DEPT ', row.get('DEPT', ''))).strip()
        gl_v = str(row.get('General Ledger Description', '')).strip()
        compound_key = f"{suffix}|{main_v}|{dept_v}|{gl_v}"
        if compound_key in refs['val_suffix_map']:
            return refs['val_suffix_map'][compound_key]
        if suffix in refs['val_suffix_map']:
            return refs['val_suffix_map'][suffix]
        return '-'

    df['Validação'] = df.apply(calc_validacao, axis=1)
    print("  ✓ Validação (PROCV Vendor# → fallback SUFFIX1)")

    # ── Expt Pymt ──
    def next_friday(dt):
        if pd.isna(dt):
            return None
        days_ahead = 4 - dt.weekday()
        if days_ahead <= 0:
            days_ahead += 7
        return dt + timedelta(days=days_ahead)

    df['Expt Pymt'] = df.apply(
        lambda r: 'PAID' if r['Status'] == 'PAID'
        else next_friday(r['Due Date']),
        axis=1
    )
    print("  ✓ Expt Pymt (próxima sexta-feira após Due Date)")

    # ── OBS ──
    df['OBS'] = None
    print("  ✓ OBS (vazio - preenchimento manual)")

    return df


# ═══════════════════════════════════════════════════════════════════════════════
# 5. GERAÇÃO DOS RESUMOS
# ═══════════════════════════════════════════════════════════════════════════════

def generate_pym_list(df):
    """Gera a PYM LIST agrupada por vendor/pagamento."""
    print("\n[4] Gerando PYM LIST...")
    unpaid = df[df['Status'] == 'UNPAID'].copy()
    if unpaid.empty:
        print("  Nenhum item UNPAID.")
        return pd.DataFrame()

    unpaid['Cod'] = unpaid['AP Type'] + unpaid['Company #'].astype(str) + unpaid['Currency']

    pym = unpaid.groupby(
        ['Cod', 'AP Type', 'Company #', 'Currency', 'Vendor #', 'Vendor Name',
         'Credit Account (Kyriba)', 'Debit Account (Kyriba)', 'Transfer Type', 'Reason'],
        as_index=False
    )['Original Amount'].sum()

    print(f"  PYM LIST: {len(pym)} linhas")
    return pym


def generate_resumo(df):
    """Gera Resumo por AP Type, Company, Currency, GL, Due Date."""
    print("  Gerando Resumo...")
    unpaid = df[df['Status'] == 'UNPAID'].copy()
    if unpaid.empty:
        return pd.DataFrame()

    resumo = unpaid.groupby(
        ['AP Type', 'Company #', 'Currency', 'GL', 'Due Date'],
        as_index=False
    ).agg({'Amount($)': 'sum'}).rename(columns={'Amount($)': 'Amount', 'Company #': 'Company'})

    resumo['Status'] = 'Approved'
    resumo['Sistema'] = 'RLM'
    resumo['COD'] = resumo['AP Type'] + '_' + resumo['Sistema'] + '_' + resumo['Currency']
    return resumo.sort_values(['AP Type', 'Company', 'Currency', 'Due Date'])


# ═══════════════════════════════════════════════════════════════════════════════
# 6. EXPORTAÇÃO COM FORMATAÇÃO
# ═══════════════════════════════════════════════════════════════════════════════

def export_to_excel(df, refs, pym, resumo, output_path):
    """Exporta tudo para Excel formatado."""
    print(f"\n[5] Exportando para: {output_path}")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='AP By GL', index=False)
        if not pym.empty:
            pym.to_excel(writer, sheet_name='PYM LIST', index=False)
        if not resumo.empty:
            resumo.to_excel(writer, sheet_name='Resumo', index=False)
        for name, key in [('Transfer Type', 'transfer_type_df'),
                          ('Currency', 'currency_df'),
                          ('Validação', 'validacao_df'),
                          ('Kyriba (C.A.)', 'kyriba_ca_df')]:
            if key in refs:
                refs[key].to_excel(writer, sheet_name=name, index=False)

    # Formatação
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(output_path)
    ws = wb['AP By GL']

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    calc_fill = PatternFill('solid', fgColor='E8F0FE')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 16

    calc_start = len(RAW_COLUMNS) + 1
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 10000),
                             min_col=calc_start, max_col=ws.max_column):
        for cell in row:
            cell.fill = calc_fill

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # Formatar PYM LIST se existir
    if 'PYM LIST' in wb.sheetnames:
        ws_pym = wb['PYM LIST']
        for cell in ws_pym[1]:
            cell.fill = PatternFill('solid', fgColor='2E75B6')
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        ws_pym.freeze_panes = 'A2'

    wb.save(output_path)

    # Stats
    paid = len(df[df['Status'] == 'PAID'])
    unpaid = len(df[df['Status'] == 'UNPAID'])
    print(f"\n{'='*60}")
    print(f"  ✅ CONCLUÍDO!")
    print(f"  Arquivo: {output_path}")
    print(f"  Linhas: {len(df)} (PAID: {paid} | UNPAID: {unpaid})")
    print(f"{'='*60}")


# ═══════════════════════════════════════════════════════════════════════════════
# 7. VALIDAÇÕES
# ═══════════════════════════════════════════════════════════════════════════════

def run_validations(df):
    """Executa validações de qualidade."""
    print("\n[VALIDAÇÕES]")

    nao_cad = {
        'Credit Account': df[df['Credit Account (Kyriba)'] == 'Não Cadastrado'],
        'Vendor Bank Country': df[df['Vendor Bank Country'] == 'Não Cadastrado'],
        'Debit Account': df[df['Debit Account (Kyriba)'] == 'Não Cadastrado'],
        'Transfer Type': df[df['Transfer Type'] == 'Não Cadastrado']
    }
    for field, subset in nao_cad.items():
        if len(subset) > 0:
            vendors = subset['Vendor #'].nunique()
            print(f"  ⚠ {field}: {len(subset)} linhas ({vendors} vendors distintos)")

    val_special = df[~df['Validação'].isin(['PAID', '-'])]
    if len(val_special) > 0:
        print("  ℹ Validações especiais:")
        for vtype, count in val_special['Validação'].value_counts().items():
            print(f"    {vtype}: {count} linhas")

    print("\n  [RESUMO UNPAID POR COMPANY/CURRENCY/AP TYPE]")
    summary = df[df['Status'] == 'UNPAID'].groupby(
        ['Company #', 'Currency', 'AP Type']
    )['Original Amount'].agg(['sum', 'count']).reset_index()
    summary.columns = ['Company', 'Currency', 'AP Type', 'Total Amount', 'Qtd Linhas']
    if not summary.empty:
        print(summary.to_string(index=False))


# ═══════════════════════════════════════════════════════════════════════════════
# 8. MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='Automação Check-in AP — FARM Global Finance',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  python automate_checkin_ap.py --bases ./Bases --ref ./CheckinAP_Ref.xlsx
  python automate_checkin_ap.py --input ./CheckinAP-260318.xlsx
  python automate_checkin_ap.py --input ./CheckinAP-260318.xlsx --output ./Resultado.xlsx
        """
    )
    parser.add_argument('--bases', type=str, help='Pasta com bases do ERP por company')
    parser.add_argument('--ref', type=str, help='Arquivo com tabelas de referência')
    parser.add_argument('--input', type=str, help='Arquivo único já consolidado')
    parser.add_argument('--output', type=str, help='Caminho de saída')
    args = parser.parse_args()

    if args.input:
        ref_path = args.input
        input_mode = 'single'
    elif args.bases and args.ref:
        ref_path = args.ref
        input_mode = 'consolidate'
    else:
        parser.error("Use --input <arquivo.xlsx> OU --bases <pasta> --ref <ref.xlsx>")

    if args.output is None:
        today = datetime.now().strftime('%y%m%d')
        args.output = f'CheckinAP_Automatizado_{today}.xlsx'

    print("=" * 60)
    print("  AUTOMAÇÃO CHECK-IN AP — FARM Global Finance")
    print(f"  {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 60)

    print("\n[1] Carregando dados brutos...")
    if input_mode == 'consolidate':
        df = load_and_consolidate_bases(args.bases)
    else:
        df = load_single_file(args.input)
    print(f"  Total: {len(df)} linhas")

    refs = load_reference_tables(ref_path)
    df = apply_transformations(df, refs)

    pym = generate_pym_list(df)
    resumo = generate_resumo(df)

    run_validations(df)
    export_to_excel(df, refs, pym, resumo, args.output)


if __name__ == '__main__':
    main()
